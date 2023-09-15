import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook

diretorio_atual = os.path.dirname(os.path.abspath(__file__))

wb = Workbook()
ws = wb.active

diretorio_xml = diretorio_atual#'C:/Users/User/Downloads/notas'

arquivos_xml = [arquivo for arquivo in os.listdir(diretorio_xml) if arquivo.endswith('.xml')]

namespaces = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

# Cabeçalhos
ws.append(['Número da Nota', 'Data de Emissão', 'Código do Produto', 'Descrição do Produto', 'NCM', 'Quantidade', 'Valor Unitário', 'Valor Total'])



for arquivo_xml in arquivos_xml:
    caminho_arquivo = os.path.join(diretorio_xml, arquivo_xml)
    tree = ET.parse(caminho_arquivo)
    root = tree.getroot()

    
    #Dados da nota
    dadosNota = root.find('.//nfe:ide', namespaces)
    numNota = dadosNota.find('nfe:nNF', namespaces).text

    #Dados do emitente
    emitente = root.find('.//nfe:emit', namespaces)
    cnpj = emitente.find('nfe:CNPJ', namespaces).text
    nome_emitente = emitente.find('nfe:xNome', namespaces).text

    data_emissao = root.find('.//nfe:ide/nfe:dhEmi', namespaces).text.split('T')[0]

    produtos = root.findall('.//nfe:det', namespaces)
    for produto in produtos:
        codigo_produto = produto.find('nfe:prod/nfe:cProd', namespaces).text
        print(codigo_produto)
        descricao_produto = produto.find('nfe:prod/nfe:xProd', namespaces).text
        ncm = produto.find('nfe:prod/nfe:NCM', namespaces).text
        quantidade = int(float(produto.find('nfe:prod/nfe:qCom', namespaces).text))
        valor_unitario = float(produto.find('nfe:prod/nfe:vUnCom', namespaces).text)
        valor_total = float(produto.find('nfe:prod/nfe:vProd', namespaces).text)

        ws.append([numNota, data_emissao, codigo_produto, descricao_produto, ncm, quantidade, valor_unitario, valor_total])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = '"R$"#,##0.00_);[Red]\-"R$"#,##0.00'

wb.save(os.path.join(diretorio_atual, 'dados.xlsx'))
# wb.save('C:/Users/User/Downloads/notas/dados.xlsx')
